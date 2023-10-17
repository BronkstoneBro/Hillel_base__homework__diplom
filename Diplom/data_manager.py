import openpyxl
from person import Person
import datetime


class DataManager:
    def __init__(self, filename="people_data.xlsx"):
        self.filename = filename
        self.load_data()

    def load_data(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(['First Name', 'Last Name', 'Middle Name', 'Birthdate', 'Deathdate', 'Gender', 'Age'])

    def save_data(self):
        self.workbook.save(self.filename)

    def is_valid_date_format(self, date_string):
        try:
            datetime.datetime.strptime(date_string, '%d.%m.%Y')
            return True
        except ValueError:
            return False

    def validate_date(self, date_string):
        if self.is_valid_date_format(date_string):
            date_obj = datetime.datetime.strptime(date_string, '%d.%m.%Y').date()
            current_date = datetime.date.today()
            if date_obj > current_date:
                return False
            return True
        else:
            print("Ошибка: Некорректный формат даты (ожидается ДД.ММ.ГГГГ).")
            return False

    def add_person(self, first_name, last_name, middle_name, birthdate, deathdate, gender):
        while not self.validate_date(birthdate):
            birthdate = input("Ошибка: Некорректная дата рождения. Введите дату рождения (в формате ДД.ММ.ГГГГ): ")

        while deathdate and not self.validate_date(deathdate):
            deathdate = input("Ошибка: Некорректная дата смерти. Введите дату смерти (в формате ДД.ММ.ГГГГ или нажмите Enter): ")

        person = Person(first_name, last_name, middle_name, birthdate, deathdate, gender)
        age = person.calculate_age()
        if age is None:
            print("Ошибка: Некорректные данные о дате рождения или смерти.")
        else:
            print(f"Возраст: {age}")
        self.sheet.append([first_name, last_name, middle_name, birthdate, deathdate, gender, age])
        self.save_data()

    def search_people(self, search_string):
        results = []
        for row in self.sheet.iter_rows(values_only=True):
            person = Person(row[0], row[1], row[2], row[3], row[4], row[5])
            if person.match_search_criteria(search_string):
                age = person.calculate_age()
                results.append(f"{person.first_name} {person.last_name}, Возраст: {age}")
        return results
