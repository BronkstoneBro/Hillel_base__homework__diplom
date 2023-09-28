import csv
from openpyxl import Workbook


workbook = Workbook()
sheet = workbook.active

with open("CSV_read.csv", 'r', newline='', encoding='utf-8') as csvfile:
    csvreader = csv.reader(csvfile)
    for row_index, row in enumerate(csvreader):
        del row[2]
        sheet.append(row)

workbook.save('Excel_read.xlsx')

new_workbook = Workbook()
new_sheet = new_workbook.active


for col in range(1, sheet.max_column + 1):
    new_row = []
    for row in range(1, sheet.max_row + 1):
        new_row.append(sheet.cell(row=row, column=col).value)
    new_sheet.append(new_row)

new_workbook.save('Excel_read90.xlsx')

print("Данные из CSV-файла сохранены в Excel_read.xlsx без столбца Age.")
print('Таблица Excel успешно перевернута на '
      '90 градусов и сохранена в Excel_read90.xlsx')
